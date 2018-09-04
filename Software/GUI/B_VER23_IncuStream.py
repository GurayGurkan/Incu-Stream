# -*- coding: utf-8 -*-
"""
@author: guray gurkan
@ Incubator Imager Project
@ B_VER 23.0
#
"""
# autofocusing added
# plate type vs. cell subgrid animation added
# e-mail option added, now optional 
# background capture added
# lens (fine focus) control added
#  XLSWrite OK
# background enhancement added


import guidata
import guiqwt
import cv2
import numpy as np
import cv2.cv as cv
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import Gui_IncuStream
import sys
import glob
import serial
import time
from datetime import datetime as dt
import datetime
import openpyxl
import openpyxl.drawing.image as Imagexl
import matplotlib
import matplotlib.mlab as mlab
import smtplib




class MainDialog(QMainWindow,Gui_IncuStream.Ui_MainWindow):
    well = {}
    well_mask = []
    #plate_type = {0:'6 Well Plate',1:'12 Well Plate',2:'24 Well Plate',3:'48 Well Plate', 4:'96 Well Plate'}
    plate_type = [['6 Well Plate',35],['12 Well Plate',22.5],['24 Well Plate',16],['48 Well Plate', 9],['96 Well Plate',6]]
    
    
    FOVx = 0.578 # mm
    FOVy = 0.362
    IM_BACKGND =[]
    subgrid_params =''
    decimation_table = [1, .75, .5, .25, .1]
    
    toggled_wellindex =[]
    list_wellinfo ={}
    plate_current = []
    well_coordinates=[]
    well_count=0
    letters = ['A','B','C','D','E','F','G','H']
    pixmap = {}
    
    repeatFLAG = False;
    repeatDuration = 1 
    repeatPeriod = 15
    repeats=1
    target_repeats = 4
    more2come = False
  
    imagelist=()
    image_path=''
    Nimages=0
    x_max=0
    y_max=0
    
    
    vid_index =0; #Camera Index
    ser=[]
    filename_glob=''
    folder =''
    cfgfolder=''
    cap_obj=[]
    serial_obj=[]
    ports=[]
    frame=[]
    gray=[]
    ssc = 1 # Single Shot cumulative counter
    
    
    count_back=False
    cycle_on = time.time()
    running=False
    
    lens_level = 0
    
    # ----------- 09.08.2018 ----- Begin
    LensCoarse_Origin = 10
    LensFine_Origin = 8
    Lens_Origin = np.array([10,8])
    Lens_Current =np.array([LensCoarse_Origin,LensFine_Origin])
    
    # ----------- 09.08.2018 ----- End
    
    
#    xls=openpyxl.Workbook()
#    XLSsetup = xls.active
#    XLSsetup.title = "Setup"
#    
#    #XLSImage = Imagexl.Image("./images/Logo.png")
#    #XLSsetup.add_image(XLSImage,'C1')
#    XLSsetup.merge_cells('A1:F6')
#    XLSsetup.merge_cells('A1:F6')
#    XLSsetup.merge_cells('A7:F7')
#    XLSsetup['A7']="Produced by IncuScope" 
#    
#    XLSstatus = xls.create_sheet(title="Events")
#    XLSstatus.append(("Date/time","Event"))
#    
#    XLScounts = xls.create_sheet(title="Counts")
    
    
   
    
    def __init__(self,parent=None): #******************************************
            super(MainDialog,self).__init__(parent)
            self.setupUi(self)
            
            
            
            ActionQuit = QAction("Quit",self)
            ActionQuit.setShortcut("CTRL+Q")
            ActionQuit.setToolTip("Quit Program")
            ActionQuit.triggered.connect(self.handle_quit)
            
            ActionAbout = QAction("About",self)
            ActionAbout.setShortcut("CTRL+A")
            ActionAbout.setToolTip("About")
            ActionAbout.triggered.connect(self.handleAbout)
            
            main_menu = self.menuBar()
            
            file_menu = main_menu.addMenu('&File')
            file_menu.addAction(ActionQuit)
            
            about_menu = main_menu.addMenu('&About')
            about_menu.addAction(ActionAbout)
            
            self.Well_OFF=QPixmap("./images/Unfilled.bmp")
            self.Well_ON=QPixmap("./images/Filled.bmp")
            
      
            ####                        COM DETECT START                      ####
            ##                                                                  ##
            ######################################################################
            
            self.DetectPorts()
            c=len(self.ports)
            
        
            if c>0:
                self.comboBox_ports.setCurrentIndex(-1)
                for a in range(c):
                    self.comboBox_ports.addItem(self.ports[a])
                self.statusbar.showMessage("Device connected...")
                self.groupBox_2.setEnabled(True)
                self.groupBox_timelapse.setEnabled(True)
                self.buttonSaveParams.setEnabled(True)
                self.pushButton_START.setEnabled(True)
                
                
            else:
                self.statusbar.showMessage('No device connected...')
                self.groupBox_2.setEnabled(True) ###
                self.groupBox_timelapse.setEnabled(False)
                self.pushButton_START.setEnabled(False)
                
            ######################################################################
                
                
            self.timerCamera = QTimer() #LiveView
            
            self.timerImaging = QTimer() # Automated
            self.timerBackcount = QTimer()
            
            
            self.timerBackcount.timeout.connect(self.print_time_left)
            ## ----- Tab 1: Setup --------------------------------------------------------
            ## ----- SUBGRID class------
            
            self.subgrid_params = SubgridParams()
            self.drawCell()
            
            self.pushButtonBCKGND.setEnabled(False)
            
            if self.folder=='':
                    try:
                        self.folder='IncuScope'
                        glob.os.chdir('C:\\')
                        glob.os.mkdir(self.folder)
                        glob.os.chdir('C:\\' + self.folder)
                        self.folder = glob.os.getcwd()
                    except WindowsError:
                        pass
                        
            glob.os.chdir('C:\IncuScope')
            self.folder=glob.os.getcwd()
            self.IM_BACKGND=cv2.imread('Corrector.jpeg')
            #pixmap = QPixmap.fromImage(self.ToQImage(self.IM_BACKGND), Qt.AutoColor)
            #pixmap = pixmap.scaled(192,108,Qt.KeepAspectRatio)
            #self.labelBCKGND.setPixmap(pixmap)
            self.pushButtonBCKGND.clicked.connect(self.updateBCKGND)
            self.pushButtonBCKGND2.clicked.connect(self.updateBCKGND2)
            
            self.sliderRadial.valueChanged.connect(self.updateRadialgrid)
            
            
            ## ----- Tab 2: Acquisition -----------------------------------------------------
            Nplates = np.size(self.plate_type,0)
            for c in range(Nplates):
                self.platetype.addItem(self.plate_type[c][0])
                
            self.platetype.currentIndexChanged.connect(self.drawCell)
            self.combo_cols.currentIndexChanged.connect(self.drawGrid)
            self.combo_rows.currentIndexChanged.connect(self.drawGrid)
            
            self.timerCamera.timeout.connect(self.Getframe)
            self.buttonSaveParams.clicked.connect(self.ConfirmSetup)
            self.pushButton_START.clicked.connect(self.StartAcquisition)
            
            self.pushButton_HALT.clicked.connect(self.cancel_all)
            
            self.groupBox_timelapse.clicked.connect(self.repeatControl)
            self.dial_duration.valueChanged.connect(self.UpdateDuration)
            self.dial_period.valueChanged.connect(self.UpdatePeriod)
            
            self.pushButton_savewellprops.clicked.connect(self.updateWellInfo)
            self.pushButton_clearwellprops.clicked.connect(self.clearWellInfo)
            self.groupBox_wellprops.setEnabled(False)
            
            
            self.UpdateDuration()
            self.UpdatePeriod()
            
            #TOOLTIPS
            self.buttonSaveParams.setToolTip("Save Parameters")
            self.pushButton_START.setToolTip("Start count")
            self.pushButton_HALT.setToolTip("Stop count")
            self.dial_duration.setToolTip("Set duration (in hours)")
            self.dial_period.setToolTip("Set period (in minutes)")
            
            ## ----- Tab 3: Live View -----------------------------------------------------            
            self.buttonStartLive.clicked.connect(self.StartLive)
            self.buttonStopLive.clicked.connect(self.StopLive)
            self.buttonSnapshot.clicked.connect(self.SingleShot)
            self.pushButton_folderselect.clicked.connect(self.SelectFolder)

            self.buttonXn.clicked.connect(self.handleStepMove)
            self.buttonXp.clicked.connect(self.handleStepMove)
            self.buttonYn.clicked.connect(self.handleStepMove)
            self.buttonYp.clicked.connect(self.handleStepMove)

            self.buttonGo.clicked.connect(self.handleGO)
            
            self.buttonMailtest.clicked.connect(self.setMail)
            
            self.connect(self.verticalSlider,  SIGNAL("valueChanged(int)"),self.Change_Bri)            
            self.connect(self.verticalSlider_2,SIGNAL("valueChanged(int)"),self.Change_Cont) 
            self.connect(self.verticalSlider_3,SIGNAL("valueChanged(int)"),self.Change_Satur) 
            self.connect(self.verticalSlider_4,SIGNAL("valueChanged(int)"),self.Change_Expos)
            self.connect(self.verticalSlider_5,SIGNAL("valueChanged(int)"),self.Change_Hue)
            self.connect(self.verticalSlider_6,SIGNAL("valueChanged(int)"),self.Change_Gain)
            self.connect(self.verticalSlider_7,SIGNAL("valueChanged(int)"),self.Change_Sharpness)
            
            self.buttonStopLive.setEnabled(False)
            self.groupBox_Focus.setEnabled(False)
            self.groupBox_CamParam.setEnabled(False)
            self.groupBox_Navigator.setEnabled(False)
            #-----------  09.08.2108 Begin
            self.buttonLensSetOrigin.clicked.connect(self.setLensOrigin)
            self.buttonLensReset.clicked.connect(self.LensToOrigin)
            self.buttonLensSavePosition.clicked.connect(self.LensSavePosition)
        
            #-----------  09.08.2108 End
            
            
            p=QPixmap(640,480)
            p.fill(Qt.blue)
            self.label_CanvasLive.setPixmap(p)
            self.label_CanvasLive.setAlignment(Qt.AlignTop | Qt.AlignLeft)
            self.setMouseTracking(True)
            self.label_CanvasLive.setMouseTracking(True)
            self.label_CanvasLive.installEventFilter(self)
            
            
            self.sliderFocuserCoarse.valueChanged.connect(self.changeCoarseLenslevel)
            self.sliderFocuserFine.valueChanged.connect(self.changeFineLenslevel)
            #self.pushButton_setPlateLEVEL.clicked.connect(self.setlenszerolevel)
            # ------------ TAB: Analysis---------------------------
            
            self.folderModel = QDirModel()
            self.folderModel.setFilter(QDir.Dirs | QDir.NoDotAndDotDot)
            
            self.folderView.setModel(self.folderModel)
            self.folderView.hideColumn(1)
            self.folderView.hideColumn(2)
            self.folderView.hideColumn(3)
            self.folderView.setRootIndex(self.folderModel.index('C:\IncuScope'))
            self.folderSelModel= self.folderView.selectionModel()
            self.folderSelModel.selectionChanged.connect(self.updateCaptureList)
            self.folderViewCWD=self.folder # initial folder
            
            self.pngSelModel = self.pngList.selectionModel()
            self.pngSelModel.currentChanged.connect(self.updateCaptureView)
            self.currentPNGprefix = ''
            self.currentPNGfolder = ''
            self.pngCycle.valueChanged.connect(self.updatePNGcycle)
            self.pngCycle.setEnabled(False)
            
            

            
    
        
            
            
    def updateCaptureList(self,inp1,inp2):
        
        path = self.folderView.selectedIndexes()[0]
        a = str(self.folderModel.filePath(path))
        pngs = returnFileList(a)
        glob.os.chdir(self.folder)
        self.pngList.clear()
        self.pngList.addItems(pngs)
        self.folderModel.refresh()
        
    
    def updateCaptureView(self,inp1,inp2):  #06.06.2016
        
        path = str(self.folderModel.filePath(self.folderView.selectedIndexes()[0]))
        png_name = str(self.pngList.currentIndex().data().toString())
        print path, png_name

        if not (png_name==None):
            png_fullname = path + '/' + png_name + '_1.png'
            pix = QPixmap(str(png_fullname))
            pix = pix.scaled(self.pngViewer.width(),self.pngViewer.height(),Qt.KeepAspectRatio)
            self.pngViewer.setPixmap(pix)
            N=repeatedName(path + '/' + png_name)
            if N>0:
                
                self.pngCycle.setEnabled(True)
                self.pngCycle.setRange(1,N)
                self.pngCycle.setMaximum=N
                self.pngCycle.setMinimum=1
                self.currentPNGprefix = path + '/' + png_name + '_'  # ../A1_
                self.label_pngCycle.setText('1')
                self.repaint()
            else:
                self.pngCycle.setEnabled(False)
            self.currentPNGfolder = path
            self.updatePNGparameters(19 + self.pngList.currentRow()) 

    def updatePNGcycle(self):  #06.06.2016
        val = self.pngCycle.value()
        self.label_pngCycle.setText(str(val))
        pix = QPixmap(self.currentPNGprefix + str(val) + '.png')
        pix = pix.scaled(self.pngViewer.width(),self.pngViewer.height(),Qt.KeepAspectRatio)
        self.pngViewer.setPixmap(pix)
        
    def updatePNGparameters(self,row):  #06.06.2016
        glob.os.chdir(self.currentPNGfolder)
        listx = glob.glob('*.xlsx')
        if len(listx):
            f = openpyxl.load_workbook(listx[0])
            s = f.get_sheet_by_name('Setup')
            output=[]
            data_list =['A','B','D','E','F']
            for c in data_list:
                if not(s[c + str(row)].value==None):
                    output.append(s[c + str(row)].value)
                else:
                    output.append('NA')
                
            glob.os.chdir(self.folder)
            self.labelPNG_Well.setText(output[0])
            self.labelPNG_Cell.setText(output[1])
            self.labelPNG_Compund.setText(output[2])
            self.labelPNG_CompConc.setText(output[3] + ' ' + output[4])
            return output
         
            
            
            
            
            
            
           
            
            
  
    ##  -------------- _init_ end -------------------------------------------------
  
    def updateBCKGND(self): # via CAMERA
        if (self.plate_current==[]):
            QMessageBox.warning(self,"IncuScope","Please select PLATE TYPE first!")
        else:
            glob.os.chdir('C:\Incuscope')
            self.folder=glob.os.getcwd()
            self.pushButtonBCKGND.setEnabled(False)
            self.pushButtonBCKGND2.setEnabled(False)
            self.backtask = updatebackgroundTASK()
            self.backtask.status_updater.connect(self.updateStatus)
            
            self.backtask.start()

    
    def updateBCKGND2(self): # via FILE
        
        self.pushButtonBCKGND.setEnabled(False)
        self.pushButtonBCKGND2.setEnabled(False)
        try:
            back_file = QFileDialog.getOpenFileName(form,"Select Background Image...","C:\IncuScope","Background files (BG*.png)")
            glob.os.chdir('C:\Incuscope')
            self.folder=glob.os.getcwd()
            self.IM_BACKGND=cv2.imread(back_file)
            pixmap = QPixmap.fromImage(self.ToQImage(self.IM_BACKGND), Qt.AutoColor)
            pixmap = pixmap.scaled(192,108,Qt.KeepAspectRatio)
            self.labelBCKGND.setPixmap(pixmap)
        except Exception:
            self.pushButtonBCKGND.setEnabled(True)
            self.pushButtonBCKGND2.setEnabled(True)
   
    def handleAbout(self):
        QMessageBox.about(self,"About Incu-Stream", "This program is developed under GNU General Public License (GNU GPL v3).\n Developed By: Guray Gurkan, PhD \n Git-Hub Page: https://github.com/GurayGurkan \n e-mail: guray_gurkan@yahoo.co.uk")
    ### ------------------------------- Tab 1 Functions ---------------------------
    def DetectPorts(self):
        
        for c in range(63):
            try:
                sub=serial.Serial(serial.device(c),timeout=3)
                sub.setTimeout(3)
                d=sub.read(4)
                
                if d=='INCU':
                    self.ports.append(serial.device(c))
                sub.close()
                time.sleep(.3)
            except serial.SerialException:
                pass
            
    def repeatControl(self):
        if self.groupBox_timelapse.isChecked():
            
            self.repeatFLAG=True;
        else:
            
            self.repeatFLAG=False;
            
 
    def generategrid(self,rows,cols):
        self.well={}
        self.list_wellinfo={}
        self.well_mask=[]
        self.focusCF=[]
        self.combo_targetwell.clear()
        H=self.groupBox_wells.height()
        
        self.Well_OFF = self.Well_OFF.scaled(H/cols,H/cols,Qt.KeepAspectRatio)
        self.Well_ON = self.Well_ON.scaled(H/cols,H/cols,Qt.KeepAspectRatio)
        
        self.well_mask=np.zeros((rows+1,cols+1),dtype=bool) #reset selections
        self.LensMatrix = np.zeros((rows+1,cols+1,2))
        self.x_max=rows;
        self.y_max=cols;
        while self.gridLayout.count():
            it = self.gridLayout.takeAt(0)
            widget = it.widget()
            widget.deleteLater()
        #REDRAW LAYOUT 
        for yi in range(cols+1):
            for xi in range(rows+1):

                    if (yi==0) and (xi==0):
                        pass
       
                    elif (xi==0) and (not(yi==0)):
                        self.well[(xi,yi)]=QLabel('')
                        self.well[(xi,yi)].setText(str(yi))
                        self.well[(xi,yi)].setAlignment(Qt.AlignBottom | Qt.AlignCenter)
                        self.gridLayout.addWidget(self.well[xi,yi],xi,yi)
                    elif (yi==0) and (not(xi==0)):
                        self.well[(xi,yi)]=QLabel('')
                        self.well[(xi,yi)].setText(self.letters[xi-1])
                        self.well[(xi,yi)].setAlignment(Qt.AlignRight | Qt.AlignCenter)
                        self.gridLayout.addWidget(self.well[xi,yi],xi,yi)
                        
                        
                    else: #x>0 and y>0
                        self.well[(xi,yi)]=QPushButton('')
                        self.well[(xi,yi)].setObjectName('%d,%d' %(xi,yi))
                        self.well[(xi,yi)].setIconSize(QSize(100,100))
                        self.well[(xi,yi)].setIcon(QIcon(self.Well_OFF))
                        self.gridLayout.addWidget(self.well[xi,yi],xi,yi)
                        self.connect(self.well[(xi,yi)],SIGNAL("clicked()"),self.ToggleWell)
                        self.list_wellinfo[(xi,yi)]=wellinfo()
                        
                        self.list_wellinfo[(xi,yi)].Well_ID=self.letters[xi-1] + '-' + str(yi)
                        self.list_wellinfo[(xi,yi)].row = xi
                        self.list_wellinfo[(xi,yi)].column = yi
                        self.combo_targetwell.addItem(self.list_wellinfo[(xi,yi)].Well_ID)
                        
        self.gridLayout.setAlignment(Qt.AlignTop | Qt.AlignLeft)

    def getImages(self,imagepath):
        glob.os.chdir(imagepath)
        image_list=glob.glob('*.jpg')
        return image_list
        
    def publishimages(self):
        for n in range(self.x_max*self.y_max):
            xi,yi=self.Count2index(n)
            fname=self.filename_glob + str(n+1) + '.jpg'
            pixmap = QPixmap(fname)
            pixmap = pixmap.scaled(600/self.y_max,700/self.x_max,Qt.KeepAspectRatio)
             
    def ToggleWell(self):
        
        index=self.well.keys()[self.well.values().index(self.sender())]
        
        if not(self.well_mask[index]):
            self.sender().setIcon(QIcon(self.Well_ON))
            self.well_mask[index]=True
            self.lineWP0.setText(self.list_wellinfo[index].Well_ID)
            
            self.groupBox_wellprops.setEnabled(True)
            
        else:
            self.sender().setIcon(QIcon(self.Well_OFF))
            self.well_mask[index]=False
            self.lineWP0.setText(self.list_wellinfo[index].Well_ID)
            self.groupBox_wellprops.setEnabled(False)
        self.toggled_wellindex = index
        self.UpdateWellInfoList()
            
            
        self.well_coordinates = np.where(self.well_mask)
        
        self.well_count = len(self.well_coordinates[0])
        #print self.well_coordinates
        
    def GenerateCommands(self,coordinates):
        xx=coordinates[0]
        yy=coordinates[1]
        command=[]
        for i in range(len(xx)):
            if yy[i]<10:
                command.append('X' + str(xx[i]) + 'Y0'+str(yy[i]))
            else:
                command.append('X' + str(xx[i]) + 'Y'+str(yy[i]))
        return command
    
    def Count2Wellname(self,count):
        
        name= self.letters[self.well_coordinates[0][count-1]-1]  + str(self.well_coordinates[1][count-1])
        return name
        
    def Count2LensCF(self,count):
        
        x = self.well_coordinates[0][count-1]
        y = self.well_coordinates[1][count-1]
        C = self.list_wellinfo[(x,y)].FocusCoarse
        F = self.list_wellinfo[(x,y)].FocusFine
        return C,F
    
    def Wellname2Message(self,name):
        
        for i in range(len(self.letters)):
            if name[0]==self.letters[i]:
                x=i+1
        y=int(name[2:])
        if y<10:
            return 'X'+str(x)+'Y0'+str(y)
        else:
            return 'X'+str(x)+'Y'+str(y)
            
    def Wellname2Coordinate(self,name):
        
        for i in range(len(self.letters)):
            if name[0]==self.letters[i]:
                x=i+1
        y=int(name[2:])
        return (x,y)
   
    def SetPlateStatus(self,status):
        itid = self.gridLayout.count()
        while itid:
            it = self.gridLayout.itemAt(itid-1)
            widget = it.widget()
            widget.setEnabled(status)
            itid = itid -1
            
   
    def ConfirmSetup(self):
        ind=self.platetype.currentIndex()
        self.plate_current=ind
        self.pushButtonBCKGND.setEnabled(True)
        if ind==0:
            self.generategrid(2,3)
            
        elif ind==1:
            self.generategrid(3,4)
            
        elif ind==2:
            self.generategrid(4,6)
            
        elif ind==3:
            self.generategrid(6,8)
            
        elif ind==4:
            self.generategrid(8,12)
        
    
        self.groupBox_wellprops.setEnabled(True)
        self.logger.clear()
        if self.Group_subgrid.isChecked():
            self.subgrid_params.flag=True
            self.subgrid_params.Ncols=self.combo_cols.currentText()
            self.subgrid_params.Nrows=self.combo_rows.currentText()
            self.subgrid_params.decimation=self.decimation_table[self.combo_decimation.currentIndex()]
        else:
            self.subgrid_params.flag=False
            
            
        
            
        
            
            
    def UpdateDuration(self):
        self.label_duration.setText(str(self.dial_duration.value()))
        self.repeatDuration=self.dial_duration.value()
        self.UpdateTargetRepeats()
        
    def UpdatePeriod(self):
        self.label_period.setText(str(self.dial_period.value()))
        self.repeatPeriod=self.dial_period.value()
        self.UpdateTargetRepeats()
        
    def UpdateTargetRepeats(self):
        self.target_repeats = (self.repeatDuration*60) / self.repeatPeriod
        self.label_repeats.setText(str(self.target_repeats))
        if self.repeatPeriod < 60:
            self.COUNT_START = dt(1900,1,1,0,self.repeatPeriod,0)
        else:
            m = self.repeatPeriod % 60
            h = self.repeatPeriod / 60
            self.COUNT_START = dt(1900,1,1,h,m,0)

    def StartAcquisition(self):
        if len(self.ports):
            if self.well_count:
                self.XLSWriteSetup()
                self.EnableControls(False)
                self.running=True
                glob.os.chdir('C:\Incuscope')
                self.folder=glob.os.getcwd()

                pctime=time.localtime()
                if self.repeatFLAG:
                    file_name= 'RM_{}_{}_{}_{}_{}_{}'.format(pctime[0],pctime[1],pctime[2],pctime[3],pctime[4],pctime[5])
                else:
                    file_name= 'SM_{}_{}_{}_{}_{}_{}'.format(pctime[0],pctime[1],pctime[2],pctime[3],pctime[4],pctime[5])
                self.addlog("Folder created.")
                glob.os.mkdir(file_name)
         
                self.filename_glob=file_name
                glob.os.chdir(self.folder + "/" + self.filename_glob)
                if self.subgrid_params.flag:
                    val = self.well_count * int(self.subgrid_params.Ncols) * int(self.subgrid_params.Nrows)
                    self.progressOverall.setMaximum(val)
                else:
                    self.progressOverall.setMaximum(self.well_count)
                self.progressOverall.setValue(0)
                
                self.task = serialTask()
                self.task.finished.connect(self.handle_taskending)
                self.task.status_updater.connect(self.updateStatus)
                self.task.progressbar_updater.connect(self.updateProgress)
                
                self.count_back=False
                
                if self.repeatFLAG:
                    if self.subgrid_params.flag:
                        val = self.well_count * int(self.subgrid_params.Ncols) * int(self.subgrid_params.Nrows) * self.target_repeats
                        self.progressCurrent.setMaximum(val)
                    else:
                        self.progressCurrent.setMaximum(self.well_count*self.target_repeats)

                    self.progressCurrent.setValue(0)
                    self.more2come=True
                    self.timerImaging.timeout.connect(self.task.start)
                    self.timer_start= self.now_str()
                    
                    self.timerImaging.setSingleShot(False)
                    self.cycle_on = time.time()
                    self.timerBackcount.start(1000)
                    self.timerImaging.start(self.repeatPeriod*60*1000) #60*1000
            
                else:
                    if self.subgrid_params.flag:
                        val = self.well_count * int(self.subgrid_params.Ncols) * int(self.subgrid_params.Nrows)
                    else:
                        val = self.well_count
                    self.progressCurrent.setMaximum(val)
                    self.progressCurrent.setValue(0)
                
                self.task.start()

                self.pushButton_HALT.setEnabled(True)
                
            else:
                QMessageBox.warning(self,"Incu-Stream","No wells selected!")
                
        else:
            self.statusbar.showMessage("No device available...")
               
    def print_time_left(self):
        if self.count_back:
            self.countStatus.setText("Next count starts in " + str(dt.strftime( self.COUNT_START + datetime.timedelta(seconds = -round(time.time()-self.cycle_on)),"%H:%M:%S")))
        
            self.countStatus.repaint()
        
            
    def Count2index(self,count):
        y_index=count //self.x_max
        if (y_index % 2)==0:
            x_index = count - self.x_max*y_index
        else:
            x_index = self.x_max*(y_index + 1) - count - 1 
        return x_index,y_index
        
        
    ### ---------------- Tab 2 Functions -----------------------------
    def handleStepMove(self):
        
        if (self.sender().objectName()=='buttonZn'):
            self.taskLive.message='Zn'
        elif (self.sender().objectName()=='buttonZp'):
            self.taskLive.message='Zp'
        elif (self.sender().objectName()=='buttonXp'):
            self.taskLive.message='xp'
        elif (self.sender().objectName()=='buttonXn'):
            self.taskLive.message='xn'
        elif (self.sender().objectName()=='buttonYn'):
            self.taskLive.message='yp'
        elif (self.sender().objectName()=='buttonYp'):
            self.taskLive.message='yn'
        elif (self.sender().objectName()=='buttonZp2'):
            self.taskLive.message='zp'
        elif (self.sender().objectName()=='buttonZn2'):
            self.taskLive.message='zn'
        self.taskLive.start()
            
    def handleGO(self):
        self.buttonGo.setEnabled(False)        
        self.timerCamera.stop()
        p=QPixmap(800,600)
        p.fill(Qt.black)
        self.label_CanvasLive.setPixmap(p)
        self.status_LiveMode.setText('Moving to target well')        
        message= self.Wellname2Message(self.combo_targetwell.currentText())
        self.repaint()
        self.taskLive.message=message
        self.taskLive.start()
        
        
        
        
    def handleLiveMovementEnd(self):
        self.buttonGo.setEnabled(True)        
        self.status_LiveMode.setText('Live Mode ON')        
        self.timerCamera.start(1000/30.)
        
    def StartLive(self):
        
        if (self.plate_current==[]):
            QMessageBox.warning(self,"Incu-Stream","Please select PLATE TYPE!")
        else:
            self.taskLive = LiveMovement('')
            self.taskLive.finished.connect(self.handleLiveMovementEnd)
            self.cap_obj = cv2.VideoCapture(self.vid_index) 
          
            self.buttonStartLive.setEnabled(False)
            self.buttonStopLive.setEnabled(True)
            self.buttonSnapshot.setEnabled(True)
            
            self.verticalSlider.setSliderPosition(self.cap_obj.get(cv.CV_CAP_PROP_BRIGHTNESS))
            self.verticalSlider_2.setSliderPosition(self.cap_obj.get(cv.CV_CAP_PROP_CONTRAST))
            self.verticalSlider_3.setSliderPosition(self.cap_obj.get(cv.CV_CAP_PROP_SATURATION))
            self.verticalSlider_4.setSliderPosition(self.cap_obj.get(cv.CV_CAP_PROP_EXPOSURE))
            self.verticalSlider_5.setSliderPosition(self.cap_obj.get(cv.CV_CAP_PROP_HUE))
            self.cap_obj.set(3,1920)
            self.cap_obj.set(4,1080)
            self.cap_obj.set(6,cv.CV_FOURCC('M', 'J', 'P', 'G') & 0xFF )
            self.verticalSlider_6.setSliderPosition(self.cap_obj.get(cv.CV_CAP_PROP_GAIN))
            self.verticalSlider_7.setSliderPosition(self.cap_obj.get(20))
            self.status_LiveMode.setText('Live Mode ON')
            self.groupBox_Focus.setEnabled(True)
            self.groupBox_CamParam.setEnabled(True)
            self.groupBox_Navigator.setEnabled(True)
            self.timerCamera.start(0)
        
            
        
    def StopLive(self):
        self.timerCamera.stop()
        self.taskLive.endtask()
        self.cap_obj.release()
        self.cap_obj={}
        
        self.buttonStartLive.setEnabled(True)
        self.buttonStopLive.setEnabled(False)
        self.groupBox_Focus.setEnabled(False)
        self.groupBox_CamParam.setEnabled(False)
        self.groupBox_Navigator.setEnabled(False)
        self.status_LiveMode.setText('Live Mode OFF')
        p = QPixmap(640,360)
        p.fill(Qt.blue)
        self.label_CanvasLive.setPixmap(p)
    
    def eventFilter(self,obj,event):
        if obj.objectName()== "label_CanvasLive" and event.type()==5:
            pix=obj.pixmap().toImage()
            val=(pix.pixel(event.x(),event.y()))
            red = QColor(val).red()
            green = QColor(val).green()
            blue = QColor(val).blue()
            self.pixel_val.setText("(" + str(event.x()) + "," + str(event.y()) + ") R: " + str(red) + " G: " + str(green) + " B:" + str(blue))
        return False    
    ##### ---------------------- CAMERA PARAMETERS -------------------
        
    def Change_Bri(self,val):
        self.label_3.setText("Brightness: " + str(val))
        self.cap_obj.set(cv.CV_CAP_PROP_BRIGHTNESS,val)
    
    def Change_Cont(self,val):
        self.label_4.setText("Contrast: " + str(val))
        self.cap_obj.set(cv.CV_CAP_PROP_BRIGHTNESS,val) 
        
    def Change_Satur(self,val):
        self.label_5.setText("Saturation: " + str(val))
        self.cap_obj.set(cv.CV_CAP_PROP_SATURATION,val)     
        
    def Change_Expos(self,val):
        self.label_6.setText("Exposure: " + str(val))
        self.cap_obj.set(cv.CV_CAP_PROP_EXPOSURE,val)
        
    def Change_Hue(self,val):
        self.label_11.setText("Hue: " + str(val))
        self.cap_obj.set(cv.CV_CAP_PROP_HUE,val)
        
    def Change_Gain(self,val):
        self.label_12.setText("Gain: " + str(val))
        self.cap_obj.set(cv.CV_CAP_PROP_GAIN,val)

    def Change_Sharpness(self,val):
        self.label_13.setText("Sharpness: " + str(val))
        self.cap_obj.set(20,val)
         
    #### ----------------------end of CAMERA PARAMETERS ----------
         
    def SelectFolder(self):
        self.folder = QFileDialog.getExistingDirectory(None, 'Select a folder:', 'C:\\', QFileDialog.ShowDirsOnly)        
        
        glob.os.chdir(str(self.folder))
        self.statusbar.showMessage('Files will be saved to "'  + self.folder + '"')
        self.status_LiveMode.setText('Files will be saved to "'  + self.folder + '"')
        time.sleep(2)
        
        self.status_LiveMode.setText("")
    
    def Getframe(self):
        ret,frame = self.cap_obj.read()
        pixmap = QPixmap.fromImage(self.ToQImage(frame), Qt.AutoColor)
        pixmap = pixmap.scaled(self.label_CanvasLive.width(), self.label_CanvasLive.height(),Qt.KeepAspectRatio)
        self.label_CanvasLive.setPixmap(pixmap)
        
    
    
        
    def SingleShot(self):
        self.buttonSnapshot.setEnabled(False)
        #cap = cv2.VideoCapture(self.vid_index)
        self.timerCamera.stop()
        time.sleep(.1)
        pctime=time.localtime()
        file_single= 'SnapShot_{}_{}_{}_{}_{}_{}_P'.format(pctime[0],pctime[1],pctime[2],pctime[3],pctime[4],pctime[5])
        flag =False
        while not flag:
            flag, frame = self.cap_obj.read()
        
        cv2.imwrite(file_single + str(self.ssc) + '.jpg',frame)
        self.ssc=self.ssc+1
        self.buttonSnapshot.setEnabled(True)
        self.status_LiveMode.setText('Saved...')
        self.timerCamera.start(1000/25);        
        self.label_CanvasLive.setMouseTracking(True)
 
    def ToQImage (self, cv_img):
        height, width, bytesPerComponent = cv_img.shape
        bytesPerLine = bytesPerComponent * width;
        cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB, cv_img)
        return QImage(cv_img.data, width, height, bytesPerLine, QImage.Format_RGB888)     
     
    
    def now_str(self):
        t = dt.now()
        return t.strftime("%d.%m.%Y %H:%M:%S")
    
    def EnableControls(self,flag):
        self.groupBox_2.setEnabled(flag)
        self.buttonSaveParams.setEnabled(flag)
        self.pushButton_START.setEnabled(flag)
        self.SetPlateStatus(flag)
        self.groupBox_timelapse.setEnabled(flag)
        
        self.pushButton_HALT.setEnabled(not(flag))
        
        
    def handle_taskending(self):
        if self.repeatFLAG:
         
            if not(self.repeats==self.target_repeats):
                form.repeats =form.repeats +1
                self.count_back=True
            else:
                self.more2come=False
                self.timerImaging.stop()
            
        if not(self.more2come):
            form.EnableControls(True)
            form.statusbar.showMessage("Finished...")
            form.addlog("Count finished.")
            form.countStatus.setText(" ")
            form.repeats=1
            self.running=False
            self.progressCurrent.setValue(0)
            self.progressOverall.setValue(0)
            self.xls.save(self.filename_glob + ".xlsx")
            
            QMessageBox.information(self, "Incu-Stream","Finished...")
            
            
            
    
    def handle_quit(self,state):
        if self.running:
            QMessageBox.warning(self,"Warning...", "Can not quit, device running...",QMessageBox.Ok)
            
        else:
            secim = QMessageBox.question(self,"Exit...", "Are you sure?",QMessageBox.Yes |  QMessageBox.No)
            if secim ==QMessageBox.Yes:
                sys.exit()
            
        
    def closeEvent(self,event):
        
        if self.running:
            QMessageBox.warning(self,"Warning...", "Can not quit, device running. Press ""Cancel"" first.",QMessageBox.Ok)
            event.ignore()
        else:
            secim = QMessageBox.question(self,"Exit...", "Do you really want to quit?",QMessageBox.Yes |  QMessageBox.No)
            if secim ==QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()
                
    def cancel_all(self):
        choice = QMessageBox.question(self,"Exit...", "Are you sure?",QMessageBox.Yes |  QMessageBox.No)
        if choice ==QMessageBox.Yes:
            self.timerImaging.stop()
            self.timerBackcount.stop()
            self.task.terminate()
            form.EnableControls(True)
            form.statusbar.showMessage("Cancelled...")
            form.addlog("Count cancelled by user.")
            form.countStatus.setText(" ")
            QMessageBox.information(self, "Incu-Stream","Cancelled...")
            form.repeats=1
            self.running=False
            self.progressCurrent.setValue(0)
            self.progressOverall.setValue(0)
            self.xls.save(self.filename_glob + ".xlsx")
     
        else:
            pass
        
    
        
    
    def updateStatus(self,text):
        self.statusbar.setVisible(True)
        self.statusbar.showMessage(text)
        
    def updateProgress(self,current,overall):
        self.progressOverall.setValue(overall)
        self.progressCurrent.setValue(current)
    
    def start_paradigm(self):
        pass
    
    def addlog(self,message):
        text = self.now_str() + "   " + message
        self.logger.addItem(text)
        self.XLSstatus.append((self.now_str(),message))
        
        
    def updateWellInfo(self):
        index = self.toggled_wellindex
        self.list_wellinfo[index].Cell_Name = self.lineWP1.text()
        self.list_wellinfo[index].Cell_Count = self.lineWP2.text()
        self.list_wellinfo[index].Compounds = self.lineWP3.text()
        self.list_wellinfo[index].CompConst = self.lineWP4.text()
        self.list_wellinfo[index].CompUnits = self.unitsCompound.currentIndex()
        QMessageBox.warning(self,"Well Info", "Parameters Saved...",QMessageBox.Ok)
        
        
    def clearWellInfo(self):
        index = self.toggled_wellindex
        self.list_wellinfo[index].Cell_Name=''
        self.list_wellinfo[index].Cell_Count=''
        self.list_wellinfo[index].Compounds=''
        self.list_wellinfo[index].CompConst=''
        self.unitsCompound.setCurrentIndex(-1)
        self.UpdateWellInfoList()
        
    def UpdateWellInfoList(self):
        index = self.toggled_wellindex
        self.lineWP1.setText(self.list_wellinfo[index].Cell_Name)
        self.lineWP2.setText(self.list_wellinfo[index].Cell_Count)
        self.lineWP3.setText(self.list_wellinfo[index].Compounds)
        self.lineWP4.setText(self.list_wellinfo[index].CompConst)
        self.unitsCompound.setCurrentIndex(self.list_wellinfo[index].CompUnits)
        
        
    def XLSWriteSetup(self):
        
        self.xls=openpyxl.Workbook()
        self.XLSsetup = self.xls.active
        self.XLSsetup.title = "Setup"
        
        #XLSImage = Imagexl.Image("./images/Logo.png")
        #XLSsetup.add_image(XLSImage,'C1')
        self.XLSsetup.merge_cells('A1:F6')
        self.XLSsetup.merge_cells('A1:F6')
        self.XLSsetup.merge_cells('A7:F7')
        self.XLSsetup['A7']="Produced by Incu-Stream" 
        
        self.XLSstatus = self.xls.create_sheet(title="Events")
        self.XLSstatus.append(("Date/time","Event"))
        
        self.XLScounts = self.xls.create_sheet(title="Counts")
            
        self.XLSsetup['A11']='Plate Type :'
        self.XLSsetup['B11']=str(self.plate_type[self.plate_current][0])
        self.XLSsetup['A12']='Count Type :'
        if self.repeatFLAG:
            self.XLSsetup['B12']='Repeated'
            self.XLSsetup['A13']='Period :'
            self.XLSsetup['B13']= str(self.repeatPeriod) + ' minutes'
            self.XLSsetup['A14']= 'Repeats :'
            self.XLSsetup['B14']= str(self.target_repeats) + ' times'
        else:
            self.XLSsetup['B12']='Single'
        
        self.XLSsetup['A16']=' '
        self.XLSsetup.append(("Well Name","Cell Name","Cell Number","Drug","Drug Concentration","units"))
        k = self.well_coordinates
        sub =np.size(self.well_coordinates[0],0)
        self.XLSsetup['A18']=' '
        for index in range(sub):
            a=k[0][index]
            b=k[1][index]
            self.XLSsetup.append((str(self.list_wellinfo[(a,b)].Well_ID),str(self.list_wellinfo[(a,b)].Cell_Name),str(self.list_wellinfo[(a,b)].Cell_Count),str(self.list_wellinfo[(a,b)].Compounds),str(self.list_wellinfo[(a,b)].CompConst),str(self.unitsCompound.itemText(self.list_wellinfo[(a,b)].CompUnits))))
    
    def drawCell(self):
        N = self.platetype.currentIndex()
        R = self.plate_type[N][1]
        FOVx_sc = 250.*self.FOVx/R
        FOVy_sc = 250.*self.FOVy/R
        gridpix = QPixmap(self.grids.width(),self.grids.height())
        gridpix.fill(QColor(230,230,230))
        painter = QPainter()
        painter.begin(gridpix)
        painter.drawEllipse(45,45,250,250)
        painter.drawLine(35,45,35,45+250)
        painter.drawLine(45,35,45+250,35)
        painter.drawText(140,20,"D: " + str(R) + " mm")
        painter.drawText(5,125,"Rows")
        x,y = center2ud(170,170,FOVx_sc,FOVy_sc)
        painter.drawRect(x,y,FOVx_sc,FOVy_sc)
        painter.end()
        self.grids.setPixmap(gridpix)
        
    def drawGrid(self):
        N = self.platetype.currentIndex()
        R = self.plate_type[N][1]
       
        FOVx_sc = 250*self.FOVx/R
        FOVy_sc = 250*self.FOVy/R
        
        self.drawCell()
        pixm = self.grids.pixmap()
        painter = QPainter()
        painter.begin(pixm)
        cartesianGrid(painter,FOVx_sc,FOVy_sc,170,170,int(self.combo_cols.currentText()),int(self.combo_rows.currentText()))
        painter.end()
        self.grids.setPixmap(pixm)
        
        
    def setMail(self):
        text = "This test mail is sent by Incu-Stream Setup..."
        sendStatus(text)
    
    def updateRadialgrid(self,val):
        N = self.platetype.currentIndex()
        
        if (N==4) & (val>1):
            val=1
        if (N==3) & (val>2):
            val=2
        if (N==2) & (val>4):
            val=4
        self.combo_cols.setCurrentIndex(val+1)
        self.combo_rows.setCurrentIndex(2*val+2)
        
    def changeCoarseLenslevel(self,level):
        
        if level>self.Lens_Current[0]:
            for c in range(level-self.Lens_Current[0]):
                self.taskLive.message='Zp'
                self.taskLive.start()
                print "Lens Up"
        elif level<self.Lens_Current[0]:
            for c in range(self.Lens_Current[0]-level):
                self.taskLive.message='Zn'
                self.taskLive.start()
                print "Lens Down"
        time.sleep(.1)
        self.Lens_Current[0] = level
        
    def changeFineLenslevel(self,level):
        
        if level>self.Lens_Current[1]:
            for c in range(level-self.Lens_Current[1]):
                self.taskLive.message='zp'
                self.taskLive.start()
                print "FineFocus: Lens Up"
        elif level<self.Lens_Current[1]:
            for c in range(self.Lens_Current[1]-level):
                self.taskLive.message='zn'
                self.taskLive.start()
                print "FineFocus: Lens Down"
        time.sleep(.1)
        self.Lens_Current[1] = level   
        print self.Lens_Current
            
        
    def setLensOrigin(self):
        self.Lens_Current = np.array([10,8])
        self.sliderFocuserCoarse.setValue(10)
        self.sliderFocuserFine.setValue(8)
        
    def LensToOrigin(self):
        self.moveLens(10,8)
        self.sliderFocuserCoarse.setValue(10)
        self.sliderFocuserFine.setValue(8)
    
    def LensSavePosition(self):
        coor = self.Wellname2Coordinate(self.combo_targetwell.currentText())
        print coor
        self.list_wellinfo[coor].FocusCoarse = self.sliderFocuserCoarse.value()
        self.list_wellinfo[coor].FocusFine = self.sliderFocuserFine.value()
        
    
    def moveLens(self,C_target,F_target):
        Csteps=C_target - self.Lens_Current[0]
        Fsteps=F_target - self.Lens_Current[1]
        
        if Csteps>0:
            command1='Zp'
        elif Csteps<0:
            command1='Zn'
        
        if Fsteps>0:
            command2='zp'
        elif Fsteps<0:
            command2='zn'
        print "C"
        
        for i in range(abs(Csteps)):
            self.taskLive.message=command1
            self.taskLive.start()
            time.sleep(.1)
            print "."
            self.taskLive.wait()
        
        print "F"
        for i in range(abs(Fsteps)):
            self.taskLive.message=command2
            self.taskLive.start()
            print "."
            time.sleep(.1)
            self.taskLive.wait()
        
        self.Lens_Current=np.array([C_target,F_target])
        print "Finished Movement"
            
            

    
    
    
    
    
    
###### END OF MAIN WINDOW CLASS ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
    
    
    
    
    
    
class updatebackgroundTASK(QThread):
    status_updater =pyqtSignal(str)
    def __init__(self, parent=None):
        super(updatebackgroundTASK, self).__init__(parent)
    def terminated(self):
        self.wait()
    def run(self):
        try:
            form.port = form.comboBox_ports.currentText()
            ser=serial.Serial(str(form.port))
            time.sleep(.1)
            
            if ser.isOpen():
                
                message="Port Opened..."
                self.status_updater.emit(message)
                ser.flush();
                time.sleep(.1)
                ser.read(4) #Read ACK and free buffer
                reading = True;
                ser.write("P" + str(form.plate_current)) #transmit type initiator
                time.sleep(.1)
                while reading:
                    inByte=ser.read();
                    if inByte=="O":
                        reading = False
                
                message="Plate type selected..."
                
                self.status_updater.emit(message)
                
                time.sleep(.1)
                
                cap = cv2.VideoCapture(form.vid_index)
                cap.set(3,1920)
                
                cap.set(4,1080)
                cap.set(6,cv.CV_FOURCC('M', 'J', 'P', 'G') & 0xFF )
                cap.set(17,5000)                  
                
                message = "Moving to B-2"
                
                self.status_updater.emit(message)
                
                ser.write("X2Y02") #send XmYn
                not_shaked=True;
                while (not_shaked):
                    inByte=ser.read()
                    if inByte=='O':
                        not_shaked=False 
                message="Capturing background..."
                
                self.status_updater.emit(message)
                ret = False
                while not(ret):
                    ret, frame = cap.read()
                    ret, frame = cap.read()
                    time.sleep(1)
                
                b1 = np.max(frame)
                white_matrix = 255*np.ones((1080,1920,3),dtype='uint8')
                
                imback = white_matrix - frame
                
                form.IM_BACKGND = imback - (255 - b1)
                
                pixmap = QPixmap.fromImage(form.ToQImage(form.IM_BACKGND), Qt.AutoColor)
                pixmap = pixmap.scaled(192,108,Qt.KeepAspectRatio)
                
                
                cv2.imwrite('BG_' + dt.now().strftime("%d%m%Y_%H%M") + ".png",form.IM_BACKGND,(cv2.cv.CV_IMWRITE_PNG_COMPRESSION,4))
                message="Background update finished..."
                
                self.status_updater.emit(message)
                form.pushButtonBCKGND.setEnabled(True)
                form.pushButtonBCKGND2.setEnabled(True)
                time.sleep(2)
                message="Ready..."
                self.status_updater.emit(message)
                
        except serial.SerialException:
            message="Error in 'backgroundTASK' ..."
            
            self.status_updater.emit(message)
                    
            return False
                    
   
        
class serialTask(QThread):
    status_updater =pyqtSignal(str)
    progressbar_updater=pyqtSignal(int,int)
    
    def __init__(self, parent=None):
        super(serialTask, self).__init__(parent)
    
    def terminated(self):
        self.wait()
    def run(self):
        try:    
                
                form.cycle_on = time.time()
                if (form.repeats>1) and form.repeatFLAG:
                    text = ("Count "+ str(form.repeats) + " of " + str(form.target_repeats) + " started.")
                    
                    
                    form.addlog(text)
                    
                    
                
                if (form.repeats == form.target_repeats):  
                    form.count_back=False
                    form.countStatus.setText("In final cycle.")
                    

                form.port = form.comboBox_ports.currentText()
                ser=serial.Serial(str(form.port))
                time.sleep(.1)
                
                if ser.isOpen():
                    try:
                        message="Port Opened..."
                        form.addlog(message)
                        self.status_updater.emit(message)
                        ser.flush();
                        time.sleep(.1)
                        ser.read(4) #Read ACK and free buffer
                        reading = True;
                        ser.write("P" + str(form.plate_current)) #transmit type initiator
                        time.sleep(.1)
                        while reading:
                            inByte=ser.read();
                            if inByte=="O":
                                reading = False
                                
                        message="Plate type selected..."
                        form.addlog(message)
                        self.status_updater.emit(message)
                        
                        time.sleep(.1)      
                        cap = cv2.VideoCapture(form.vid_index)
                        cap.set(3,1920)
                        cap.set(4,1080)
                        cap.set(6,cv.CV_FOURCC('M', 'J', 'P', 'G') & 0xFF )
                        cap.set(13,0)
                        cap.set(17,5000)
                        
                        
                        
                        message="Starting capture..."
                        form.addlog(message)
                        self.status_updater.emit(message) 
                        
                        mess=form.GenerateCommands(form.well_coordinates)
                        
                        local_counter=1
                        
                        if form.subgrid_params.flag:
                            
                            Nc = int(form.subgrid_params.Ncols)
                            Nr = int(form.subgrid_params.Nrows)
                            fs = form.subgrid_params.decimation 
                            IMROWS =  Nr*1080*fs
                            IMCOLS =  1920*fs
                            reading = True;
                            
                            ser.write("r" + str(form.subgrid_params.Nrows) + "c" + str(form.subgrid_params.Ncols))
                            while reading:
                                inByte=ser.read()
                                if inByte=="O":
                                    reading = False
                            grid_counter=0    
                            message="Subgrid set to (" + str(Nr) + ',' + str(Nc) + ")"
                            form.addlog(message)
                            self.status_updater.emit(message)
                        
                        
                        
                        for i in mess: # for each well
                            name = form.Count2Wellname(local_counter)
                            message = "Moving to " + name
                            form.addlog(message)
                            self.status_updater.emit(message)
                            
                            print i
                            ser.write(i) #send XmYn
                            time.sleep(.2)
                            waitMessage('O',ser)
                       
                            # At Target Well NOW!
                       
                            message = name + ": Adjusting Preset Focus..."
                            print message
                            form.addlog(message)
                            self.status_updater.emit(message)
                            LensC, LensF= form.Count2LensCF(local_counter)
                                                       
                            #Lens Movement
                            Csteps= LensC - form.Lens_Current[0]
                            Fsteps= LensF - form.Lens_Current[1]
                            
                            if Csteps>0:
                                command1='Zp'
                            elif Csteps<0:
                                command1='Zn'
                            
                            if Fsteps>0:
                                command2='zp'
                            elif Fsteps<0:
                                command2='zn'
                            
                            for i in range(abs(Csteps)):
                                ser.write(command1)
                                waitMessage('O',ser)
                            
                            for i in range(abs(Fsteps)):
                                ser.write(command2)
                                waitMessage('O',ser)
                            
                            form.Lens_Current=np.array([LensC,LensF])
                            time.sleep(.1)
                            ret = False
                            while not(ret):
                                ret, init_frame = cap.read()
                                ret, init_frame = cap.read()
                            time.sleep(.2)    
                            
                            
                            if form.subgrid_params.flag:
                                subframe = np.zeros((1080*fs,1920*fs,3),dtype='uint8')
                                

                                ser.write('G')
                                time.sleep(.1)
                                message = name + ": Started sub-grid capturing..."
                                print message
                                self.status_updater.emit(message)
                                form.addlog(message)
                                mainframe = np.zeros((IMROWS,0,3),dtype='uint8');
        
                                for grid_cols in range(Nc):
                                    for grid_rows in range(Nr):
                                        not_shaked=True;
                                        print "Entering", grid_rows+1, grid_cols+1
                                        while (not_shaked):
                                            inByte=ser.read()
                                            if (inByte=='C') | (inByte=='W'):
                                                not_shaked=False
                                        if inByte == 'C':
                                            if grid_rows==0:
                                                subframe.fill(0)
                                                Vframe = np.zeros((0,IMCOLS,3),dtype='uint8')

                                            time.sleep(.2) #mechanical settle duration                                     #24.05.2016

#                                            N_AVERAGE=2
#                                            for AVER in range(N_AVERAGE): #average image for CMOS sensor denoising
#                                                ret = False
#                                                while not(ret):
#                                                    ret, frame = cap.read()
#                                                    ret, frame = cap.read()
#                                                print AVER, np.max(subframe)
#                                                subframe=cv2.addWeighted(subframe.astype('float32'),1,cv2.resize(frame.astype('float32'),None,None,fs,fs),0.5,1)
                                            ret = False
                                            while not(ret):
                                                ret, subframe = cap.read()
                                                ret, subframe = cap.read()
                                                
                                            #subframe = cv2.resize(subframe.astype('float32'),None,None,fs,fs)
                                            subframe = cv2.resize(subframe,None,None,fs,fs)

                                            
                                            Vframe = np.vstack((subframe,Vframe)) # (Scan Type 3)
                                            
#Scan type 1
#                                            if (grid_cols % 2):
#                                                Vframe = np.vstack((subframe,Vframe))
#                                                
#                                            else:                                  
#                                                Vframe = np.vstack((Vframe,subframe))
                 
                                            ser.write('O')
                                            message = name + " subgrid, Row: " + str(grid_rows+1)+ ", Column: " + str(grid_cols+1)
                                            print message
                                            self.status_updater.emit(message)
                                            form.addlog(message)
                                            time.sleep(.1)
                                            grid_counter +=1
                                            self.overall=(local_counter-1) + grid_counter + (form.repeats-1)*(form.well_count * int(form.subgrid_params.Ncols) * int(form.subgrid_params.Ncols))
                                            self.progressbar_updater.emit((local_counter-1) + grid_counter,self.overall)
                                 
                                    mainframe = np.hstack((Vframe,mainframe))
                                    del Vframe
                                    print "Column end"
                                    
                                not_shaked=True;
                                while (not_shaked):
                                    inByte=ser.read()
                                    if inByte=='W':
                                        not_shaked=False
                                gray = mainframe
                                
                                
                            else: # non-subgrid
                                print "No subgrid Capturing"
                                gray = init_frame
                                
                                self.overall=local_counter  + (form.repeats-1)*form.well_count
                                
                                
                                #self.progressbar_updater.emit(local_counter,self.overall)
                            local_counter = local_counter +1
                          
                            message = name + ": Saving Image..." 
                            form.addlog(message)
                            self.status_updater.emit(message)
                            #gray = cv2.cvtColor(mainframe, cv2.COLOR_BGR2GRAY)
                            
                            cv2.putText(gray,"Incu-Stream", (5,25), cv2.FONT_HERSHEY_PLAIN, 2, 255)
                            cv2.putText(gray,form.plate_type[form.plate_current][0], (5,60), cv2.FONT_HERSHEY_PLAIN, 2, 255)
                            cv2.putText(gray, "Well: " + name, (5,95), cv2.FONT_HERSHEY_PLAIN, 2, 255)
                            if form.repeatFLAG:
                                cv2.putText(gray, "Cycle: " + str(form.repeats) , (5,130), cv2.FONT_HERSHEY_PLAIN, 2, 255)
                                                        
                            cv2.imwrite(name + "_" + str(form.repeats)  + ".png" ,gray,(cv2.cv.CV_IMWRITE_PNG_COMPRESSION,4))
                            time.sleep(.1)

                        ser.write('F') # Finish
                        
                        if form.repeatFLAG:
                            message = "Cycle " + str(form.repeats) + " of " + str(form.target_repeats) + " finished." 
                            self.status_updater.emit(message)
                            form.addlog(message)
                            sendStatus(message)
                        time.sleep(.1)
                        ser.close()
                        cap.release()
                        cv2.destroyAllWindows()
                      
                        
                    except Exception:
                        message="Error..."
                        
                        sendStatus("Error occured")
                        self.status_updater.emit(self.message)
                        form.addlog(message)
                        ser.close()
                        cap.release()
                        cv2.destroyAllWindows()
                        return False
                else:
                    self.message="Port can not be opened..."
                    self.status_updater.emit(self.message)
                    sendStatus(self.message)
                    return False
                
        except serial.SerialException:
                self.message="Port can not be opened..."
                self.status_updater.emit(self.message)
                sendStatus(self.message)
                
    
        
class LiveMovement(QThread):
    def __init__(self,message,parent=None):
        super(LiveMovement, self).__init__(parent)
        
        self.message=str(message)
        form.port = form.comboBox_ports.currentText()
        self.serial_obj = serial.Serial(str(form.port))
        time.sleep(.1)
        
        self.serial_obj.read(4)
        self.serial_obj.write("P" + str(form.plate_current)) #transmit type initiator
        time.sleep(.1)
        
        reading =True
        while reading:
            inByte=self.serial_obj.read()
            
            if inByte=="O":
                reading=False
                
                
    def terminated(self):
                
        self.wait()
        
    def run(self):
        
        self.serial_obj.write(self.message)
        done=False
        
        while (not done):
            
            inByte=self.serial_obj.read()
            if inByte=='O' or inByte=='C':
                done=True
        self.finished.emit()
        return
    def endtask(self):
        self.serial_obj.write('F')
        self.serial_obj.close()
        self.terminate()

        
            
            
         

class wellinfo(object):
        def __init__(self):
                self.Well_ID=''
                self.row = ''
                self.column=''
                self.Cell_Name=''
                self.Cell_Count=''
                self.Compounds=''
                self.CompConst=''
                self.CompUnits=-1
                self.FocusCoarse = 10
                self.FocusFine = 8

def AutoFocus(max_steps,vid,port):
    
    a=np.zeros((max_steps+1))
        
    ret = False
    while not(ret):
        ret, frame = vid.read()
    c3,x3=np.histogram(frame,range=(0,255),bins=256)
    
    a[0]=np.min(mlab.find(c3>10))

    min_val=a[0]
    
    port.write('Zp')
    time.sleep(.1)
    done = False
    while not done:
        if port.read() == 'O':
            done = True
        
    lastdir = True
    
    for c in range(max_steps):
        r=False
        while not r:
            r,f = vid.read()
            
        
        #a[c+1] = np.var(f)
        c3,x3=np.histogram(f,range=(0,255),bins=256)
        a[c+1]=np.min(mlab.find(c3>10))
        if a[c+1]< min_val:
           nextdir = lastdir
           min_val=a[c+1]
           frame = f
        else:
            nextdir = not(lastdir)
        
        if nextdir:
            port.write('Zp')
            
        else:
            port.write('Zn')
            
        done = False
        while not done:
            if port.read() == 'O':
                done = True
        time.sleep(.1)
        lastdir = nextdir
    
    return True, a, frame

class SubgridParams(object):
    def __init__(self):
        self.Ncols=""
        self.Nrows=""
        self.decimation=0
        self.flag=False
        
def center2ud(xc,yc,W,H):
    x_ud = xc- W/2.0
    y_ud = yc- H/2.0
    return x_ud, y_ud
    
def cartesianGrid(pic,W,H,Ox,Oy,C,R):
    rangeX = range(-C,C+1,2)
    rangeY = range(-R,R+1,2)
        
    for cols in range(C+1):
        pic.drawLine(Ox+rangeX[cols]*W/2.0, Oy+rangeY[0]*H/2 ,Ox+rangeX[cols]*W/2.0,Oy+rangeY[-1]*H/2)
    for rows in range(R+1):
        pic.drawLine(Ox+rangeX[0]*W/2.0, Oy+rangeY[rows]*H/2 ,Ox+rangeX[-1]*W/2.0,Oy+rangeY[rows]*H/2)
    

def sendStatus(msg):
    if form.groupBox_Mail.isChecked():
        fromaddr = form.lineSender.text().toLocal8Bit().data()
        toaddrs  = {form.lineRecep1.text().toLocal8Bit().data(),form.lineRecep2.text().toLocal8Bit().data(),form.lineRecep3.text().toLocal8Bit().data()}
        msg = 'Subject: %s\n\n%s' % ('Incu-Stream Status', msg)
        username = form.lineSender.text().toLocal8Bit().data()
        password = form.linePass.text().toLocal8Bit().data()
        print "ok-1579"
        server = smtplib.SMTP(form.lineSMTP.text().toLocal8Bit().data())  #'smtp.live.com:587'
        server.ehlo()
        server.starttls()
        server.login(username,password)
        server.sendmail(fromaddr, toaddrs, msg)
        server.quit()

def makehomogen(inp,aspect,offset):
    inp = inp.astype('float')
    inp = inp + 255 - np.max(inp)
    if (aspect == 0):
        p_rows = 9
        p_cols = 16
    else:
        p_rows = 3
        p_cols =4
    size_inp = inp.shape
    col_ind = np.array(range(p_cols))
    row_ind = np.array(range(p_rows))
    
    dx = size_inp[1] / p_cols
    x = col_ind * dx + dx/2
    
    dy = size_inp[0] / p_rows
    y = row_ind * dy + dy/2
    
    back_gnd = np.zeros((p_rows,p_cols,size_inp[2]))
    c = 0 
    
    x=x.astype('int')
    y=y.astype('int')
    for m in x:
        r=0
        for n in y:
            #back_gnd[r][c][0:3]=inp[n][m][:]
            interval = np.array(range(-4,5))
            
            
            back_gnd[r][c][0]=np.mean(inp[interval+n,interval+m,0])
            
            
            back_gnd[r][c][1]=np.mean(inp[interval+n,interval+m,1])
            back_gnd[r][c][2]=np.mean(inp[interval+n,interval+m,2])
            r = r+1
        c=c+1
    
    back_gnd2 =  cv2.resize(back_gnd,(size_inp[1],size_inp[0]))
    
    final = cv2.subtract(inp,back_gnd2) + offset
    final = final.astype('uint8')
    return final

def uniqueList(inp):
    output=[]
    
    for c in range(np.size(inp)):
        if inp[c][:2] not in output:
            output.append(inp[c][:2])
    return output
    
def returnFileList(path):
    glob.os.chdir(str(path))
    pnglist = glob.glob('*.png')
    lastlist = uniqueList(pnglist)
    return lastlist
    
def repeatedName(fname):
    flist = glob.glob(str(fname) + '*.png')
    
    
    if len(flist)>1:
        return len(flist)
    else:
        return 0

def focus_measure(inp):
       
    image = inp.astype('float64')
    im_size = image.shape
    Nr = im_size[0]
    Nc = im_size[1]
    roi = image[(Nr/2 - 300):(Nr/2 + 300), (Nc/2 - 300):(Nc/2 + 300),:]
    edgeX = cv2.Sobel(roi, cv2.CV_64F, 1, 0)
    edgeY = cv2.Sobel(roi, cv2.CV_64F, 0, 1)
    return np.mean(edgeX*edgeX + edgeY*edgeY) 

def AutoFocus2(max_steps,vid,port):
  
    ret = False
    while not(ret):
        ret, frame = vid.read()
        ret, frame = vid.read()
    f0 = focus_measure(frame)
    print "Focus Measure:",f0  
    fine_img = frame
    steps = 0
    
    for c in range(max_steps):
        port.write('Zp')
        done = False
        while not done:
            if port.read() == 'O':
                done = True
        ret = False
        while not(ret):
            ret, frame = vid.read()
            ret, frame = vid.read()
        f1=focus_measure(frame)
        if f1>f0:
            f0 = f1
            fine_img = frame
            steps = c+1
            print f0, steps
    for c in range(max_steps):
        port.write('Zn')
        done = False
        while not done:
            if port.read() == 'O':
                done = True
    fine_img = fine_img.astype('uint8')
  
    return fine_img, steps

    
def waitMessage(char,serial):
    not_shaked=True;
    while (not_shaked):
        inByte=serial.read()
        if inByte==char:
            not_shaked=False
    
        
            
        
        
    
    

app=QApplication(sys.argv)
form=MainDialog()
form.show()
app.exec_()




